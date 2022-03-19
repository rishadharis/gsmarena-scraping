from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
import openpyxl
import time
import subprocess
import csv
# Set Chromedriver Option
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--start-maximized")
driver = webdriver.Chrome(executable_path='./chromedriver',options=chrome_options)
driver.get("https://www.gsmarena.com/samsung_galaxy_grand_prime-6708.php")

wb = openpyxl.load_workbook("Check DEVICE.xlsx")
ws = wb[wb.sheetnames[0]]
tername = []
brand = []

i = 0
for col in ws['A']:
    if(i<=1):
        i+=1
        continue
    tername.append(str(col.value))
with open('result.csv', 'w', newline='') as file:
    writer = csv.writer(file)
    writer.writerow(["TER_MODEL_NAME", "GSMArena Technology", "URL"])
    for a in range(len(tername)):
        if "(" in tername[a]:
            cari = tername[a].split("(")[0]
            driver.find_element(By.ID, "topsearch-text").clear()
            driver.find_element(By.ID, "topsearch-text").click()
            driver.find_element(By.ID, "topsearch-text").send_keys(cari)
            driver.find_element(By.XPATH, "//*[@id='topsearch']/div[2]/a[1]").click()
            result = driver.find_elements_by_xpath("//*[@id='review-body']/div/ul/li/a")
            length=len(result)
            if not(length==1 or length==2):
                print(a+2)
                continue
                time.sleep(3)
            try:
                check = driver.find_element(By.XPATH, "//*[@id='review-body']/div/div")
            except NoSuchElementException:
                time.sleep(1)
                result[0].click()
                techno = driver.find_element(By.XPATH, "//*[@id='specs-list']/table[1]/tbody/tr[1]/td[2]/a").text
                data = [ws['A'+str(a+3)].value, techno, driver.current_url]
                print(data)
                writer.writerow(data)
            else:
                print(a+2)     
        elif "/" in tername[a]:
            cari = tername[a].split("/")[0]
            driver.find_element(By.ID, "topsearch-text").clear()
            driver.find_element(By.ID, "topsearch-text").click()
            driver.find_element(By.ID, "topsearch-text").send_keys(cari)
            driver.find_element(By.XPATH, "//*[@id='topsearch']/div[2]/a[1]").click()
            result = driver.find_elements_by_xpath("//*[@id='review-body']/div/ul/li/a")
            length=len(result)
            if not(length==1 or length==2):
                print(a+2)
                continue
                time.sleep(3)
            try:
                check = driver.find_element(By.XPATH, "//*[@id='review-body']/div/div")
            except NoSuchElementException:
                time.sleep(1)
                result[0].click()
                techno = driver.find_element(By.XPATH, "//*[@id='specs-list']/table[1]/tbody/tr[1]/td[2]/a").text
                data = [ws['A'+str(a+3)].value, techno, driver.current_url]
                print(data)
                writer.writerow(data)
            else:
                print(a+3)
                continue
                time.sleep(3)
        else:
            cari = tername[a]
            driver.find_element(By.ID, "topsearch-text").clear()
            driver.find_element(By.ID, "topsearch-text").click()
            driver.find_element(By.ID, "topsearch-text").send_keys(cari)
            driver.find_element(By.XPATH, "//*[@id='topsearch']/div[2]/a[1]").click()
            result = driver.find_elements_by_xpath("//*[@id='review-body']/div/ul/li/a")
            length=len(result)
            if not(length==1 or length==2):
                print(a+2)
                continue
                time.sleep(3)
            try:
                check = driver.find_element(By.XPATH, "//*[@id='review-body']/div/div")
            except NoSuchElementException:
                time.sleep(1)
                result[0].click()
                techno = driver.find_element(By.XPATH, "//*[@id='specs-list']/table[1]/tbody/tr[1]/td[2]/a").text
                data = [ws['A'+str(a+3)].value, techno, driver.current_url]
                print(data)
                writer.writerow(data)
            else:
                print(a+2)
                continue
            time.sleep(3)
        time.sleep(3)

